from openpyxl.descriptors import Typed

from openpyxl_templates.style import StyleSet, SheetStyleMixin


class WorkbookTemplate(SheetStyleMixin):
    sheets = None
    active_sheet = Typed(expected_class=str, allow_none=True)

    styles = Typed(expeced_class=StyleSet)

    def __init__(self, workbook, sheets=None, active_sheet=None, styles=None, **style_keys):
        super().__init__(**style_keys)

        self.workbook = workbook
        self.sheets = sheets or self.sheets or []
        self.active_sheet = active_sheet or self.active_sheet

        self.styles = styles or self.styles or StyleSet()

        self._sheet_map = {sheet.name: sheet for sheet in self.sheets}

        for sheet in self.sheets:
            sheet.inherit_styles(self)

    def write_sheet(self, name, objects):
        excel_sheet = self._sheet_map[name]
        worksheet = self.get_or_create_sheet(excel_sheet)
        excel_sheet.write(worksheet, self.styles, objects)

        self.update_active_sheet()

    def read_rows(self, name):
        excel_sheet = self._sheet_map[name]
        worksheet = self.get_or_create_sheet(excel_sheet)
        return excel_sheet.read_rows(worksheet)

    def get_or_create_sheet(self, excel_sheet):
        name = excel_sheet.name
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        return self.workbook.create_sheet(excel_sheet.name)

    def update_active_sheet(self):
        for index, sheetname in enumerate(self.workbook.sheetnames):
            if sheetname == self.active_sheet:
                self.workbook.active = index
                return




    # def _hash_style_without_name(self, style):
    #     fields = []
    #     for attr in NamedStyle.__elements__ + ("number_format",):
    #         val = getattr(style, attr)
    #         if isinstance(val, list):
    #             val = tuple(val)
    #         fields.append(val)
    #
    #     return hash(tuple(fields))
