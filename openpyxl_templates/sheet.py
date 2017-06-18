from collections import Counter
from itertools import chain
from types import FunctionType

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation

from openpyxl_templates.exceptions import BlankNotAllowed, OpenpyxlTemplateException
from openpyxl_templates.style import StyleSet, StandardStyleSet
from openpyxl_templates.utils import Typed, OrderedType

MAX_COLUMN_INDEX = column_index_from_string("XFD")

class TemplatedWorkbookNotSet(OpenpyxlTemplateException):
    def __init__(self, templated_sheet):
        super().__init__(
            "The sheet '%s' has no assosiated workbook. This should be done automatically by the TemplatedWorkbook."
            % templated_sheet.sheetname
        )

class WorksheetDoesNotExist(OpenpyxlTemplateException):
    def __init__(self, templated_sheet):
        super().__init__(
            "The workbook has no sheet '%s'." % templated_sheet.sheetname
        )

class TemplatedSheet(metaclass=OrderedType):
    sheetname = Typed("sheetname", expected_type=str)
    active = Typed("active", expected_type=bool, value=False)
    _workbook = None

    # order = ... # TODO: Add ordering to sheets either through declaration on workbook or here

    def __init__(self, sheetname=None, active=None):
        self.sheetname = sheetname or self.sheetname
        self.active = active if active is not None else self.active

    @property
    def exists(self):
        return self.sheetname in self.workbook

    @property
    def worksheet(self):
        if not self.exists:
            self.workbook.create_sheet(self.sheetname)

        return self.workbook[self.sheetname]

    @property
    def workbook(self):
        if not self._workbook:
            raise TemplatedWorkbookNotSet(self)
        return self._workbook

    @workbook.setter
    def workbook(self, workbook):
        self._workbook = workbook

    @property
    def sheet_index(self):
        try:
            return self.workbook.sheetnames.index(self.sheetname)
        except ValueError:
            raise WorksheetDoesNotExist(self)


    def write(self, *args, overwrite=True, **kwargs):
        raise NotImplemented()
        # 'self.sheet_template.write(self.worksheet, self.templated_workbook.styles, data)

    def read(self, exception_policy):
        raise NotImplemented()

    def remove(self):
        if self.exists:
            del self.workbook[self.sheetname]

    def activate(self):
        self.workbook.active = self.sheet_index


class ColumnIndexNotSet(OpenpyxlTemplateException):
    def __init__(self, column):
        super().__init__(
            "Column index not set for column '%s'. This should be done automatically by the TableSheet." % column
        )


DEFAULT_COLUMN_WIDTH = 8.43


class TableColumn:
    setter = Typed("setter", expected_type=FunctionType, allow_none=True)
    getter = Typed("getter", expected_type=FunctionType, allow_none=True)
    _column_index = None

    # Rendering properties
    _header = Typed("header", expected_type=str, allow_none=True)
    width = Typed("width", expected_types=(int, float), value=8.43)
    hidden = Typed("hidden", expected_type=bool, value=False)
    group = Typed("group", expected_type=bool, value=False)
    data_validation = Typed("data_validation", expected_type=DataValidation, allow_none=True)

    # Reading/writing properties
    default_value = None  # internal value
    allow_blank = Typed("allow_blank", expected_type=bool, value=True)

    header_style = Typed("header_style", expected_type=str, value="Header")
    row_style = Typed("row_style", expected_type=str, value="Row")

    BLANK_VALUES = (None, "")

    def __init__(self, object_attr=None, getter=None, setter=None, header=None, width=None, hidden=None, group=None,
                 data_validation=None, default_value=None, allow_blank=None):
        self._header = header if header is not None else self._header
        self.width = width if width is not None else self.width
        self.hidden = hidden if hidden is not None else self.hidden
        self.group = group if group is not None else self.group
        self.data_validation = data_validation if data_validation is not None else self.data_validation

        self.default_value = default_value if default_value is not None else self.default_value
        self.allow_blank = allow_blank if allow_blank is not None else self.allow_blank

        if object_attr:
            self.getter = lambda obj: getattr(obj, object_attr)
            self.setter = lambda obj, value: setattr(obj, object_attr, value)

        self.getter = getter if getter is not None else self.getter
        self.setter = setter if setter is not None else self.setter

    def get_value(self, obj):
        return self.getter(obj)

    def set_value(self, obj, value):
        self.setter(obj, value)

    def to_excel(self, value):
        return value

    def from_excel(self, cell):
        return cell.value

    def to_excel_with_blank_check(self, value):
        if value is None:
            if self.allow_blank:
                return None
            raise BlankNotAllowed()
        return self.to_excel(value)

    def from_excel_with_blank_check(self, cell):
        if cell.value in self.BLANK_VALUES:
            if not self.allow_blank:
                raise BlankNotAllowed(cell=cell)
            return self.default_value
        return self.from_excel(cell)

    def prepare_worksheet(self, worksheet):
        if self.data_validation:
            worksheet.add_data_validation(self.data_validation)

    def create_header(self, worksheet):
        header = WriteOnlyCell(ws=worksheet, value=self.header)
        header.style = self.header_style
        return header

    def create_cell(self, worksheet, obj=None):
        cell =  WriteOnlyCell(
            worksheet,
            value=self.to_excel_with_blank_check(
                self.get_value(obj) if obj is not None else self.default_value
            )
        )
        cell.style = self.row_style
        return cell

    def post_process_cell(self, worksheet, cell):
        if self.data_validation:
            self.data_validation.add(cell)

    def post_process_worksheet(self, worksheet):
        column_dimension = worksheet.column_dimensions[self.column_letter]
        column_dimension.hidden = self.hidden
        column_dimension.width = self.width

    @property
    def header(self):
        return self._header or "Column%d" % self.column_index

    @property
    def column_index(self):
        if self._column_index is None:
            raise ColumnIndexNotSet(self)
        return self._column_index

    @column_index.setter
    def column_index(self, value):
        self._column_index = value

    @property
    def column_letter(self):
        return get_column_letter(self.column_index)


class ColumnHeadersNotUnique(OpenpyxlTemplateException):
    def __init__(self, columns):
        counter = Counter(column.header for column in columns)
        super().__init__("headers '%s' has been declared more then once in the same TableSheet" % tuple(
            header
            for (header, count)
            in counter.items()
            if count > 1
        ))


class TempleteStyleNotFound(OpenpyxlTemplateException):
    def __init__(self, missing_style_name, style_set):
        super().__init__(
            "The style '%s' has not been declared. Avaliable styles are: %s)"
            % (missing_style_name, style_set.names)
        )


class NoTableColumns(OpenpyxlTemplateException):
    def __init__(self, table_sheet):
        super().__init__(
            "The TableSheet '%s' has no columns. Declare atleast one."
            % table_sheet.sheetname
        )


class TableSheet(TemplatedSheet):
    item_class = TableColumn

    title_style = Typed("title_style", expected_type=str, value="Title")
    description_style = Typed("description_style", expected_type=str, value="Description")

    freeze_header = Typed("freeze_header", expected_type=bool, value=True)
    hide_excess_columns = Typed("hide_excess_columns", expected_type=bool, value=True)

    def __init__(self, sheetname=None, active=None):
        super().__init__(sheetname=sheetname, active=active)

        self.columns = list(self._items.values())

        for index, column in enumerate(self.columns):
            column.column_index = index + 1  # Start as 1

        self._validate()

    def _validate(self):
        self._check_atleast_one_column()
        self._check_unique_column_headers()

    def _check_atleast_one_column(self):
        if not self.columns:
            raise NoTableColumns(self)

    def _check_unique_column_headers(self):
        if len(set(column.header for column in self.columns)) < len(self.columns):
            raise ColumnHeadersNotUnique(self.columns)

    def write(self, title=None, description=None, objects=None):
        worksheet = self.worksheet

        self.prepare_worksheet(worksheet)
        self.write_title(worksheet, title)
        self.write_description(worksheet, description)
        self.write_headers(worksheet)
        self.write_rows(worksheet, objects)
        self.post_process_worksheet(worksheet)

    def prepare_worksheet(self, worksheet):
        for column in self.columns:
            column.prepare_worksheet(worksheet)

        # Register styles
        style_names = set(chain(
            (self.title_style, self.description_style),
            *((column.row_style, column.header_style) for column in self.columns)
        ))

        existing_names = set(self.workbook.named_styles)

        for name in style_names:
            if name in existing_names:
                continue

            if name not in self.workbook.template_styles:
                raise TempleteStyleNotFound(name, self.workbook.template_styles)

            self.workbook.add_named_style(self.workbook.template_styles[name])

    def write_title(self, worksheet, title=None):
        if not title:
            return

        title = WriteOnlyCell(ws=worksheet, value=title)
        title.style = self.title_style

        worksheet.append((title,))

    def write_description(self, worksheet, description=None):
        if not description:
            return

        description = WriteOnlyCell(ws=worksheet, value=description)
        description.style = self.description_style

        worksheet.append((description,))

    def write_headers(self, worksheet):
        self.worksheet.append(
            column.create_header(worksheet)
            for column in self.columns
        )

    def write_rows(self, worksheet, objects=None):
        for obj in objects:
            cells = tuple(column.create_cell(worksheet, obj) for column in self.columns)
            worksheet.append(cells)

            for cell, column in zip(cells, self.columns):
                column.post_process_cell(worksheet, cell)

    def post_process_worksheet(self, worksheet):
        for column in self.columns:
            column.post_process_worksheet(worksheet)

        if self.active:
            self.activate()

    def read(self, exception_policy):
        pass


class TemplatedWorkbook(Workbook, metaclass=OrderedType):
    item_class = TemplatedSheet

    templated_sheets = None
    template_styles = Typed("template_styles", expected_type=StyleSet)

    def __new__(cls, *args, file=None, **kwargs):
        if file:
            return load_workbook(file)
        return super().__new__(cls)

    def __init__(self, template_styles=None):
        super().__init__()

        self.templated_sheets = list(self._items.values())
        for sheet in self.templated_sheets:
            sheet.workbook = self
        self.template_styles = template_styles or self.template_styles or StandardStyleSet()

    def remove_all_sheets(self):
        for sheetname in self.sheetnames:
            del self[sheetname]
