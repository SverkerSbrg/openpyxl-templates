import re
from collections import Counter, namedtuple
from collections import OrderedDict
from enum import Enum
from itertools import chain, repeat, groupby

from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from openpyxl_templates.exceptions import CellExceptions, RowExceptions, SheetException, CellException
from openpyxl_templates.table_sheet.columns import TableColumn
from openpyxl_templates.templated_sheet import TemplatedWorksheet
from openpyxl_templates.utils import Typed, MAX_COLUMN_INDEX


class TableSheetException(SheetException):
    pass


class IgnoreRow(Exception):
    pass


class ColumnHeadersNotUnique(TableSheetException):
    def __init__(self, columns):
        counter = Counter(column.header for column in columns)
        super(ColumnHeadersNotUnique, self).__init__("headers '%s' has been declared more then once in the same TableSheet" % str(tuple(
            header
            for (header, count)
            in counter.items()
            if count > 1
        )))


class TempleteStyleNotFound(TableSheetException):
    def __init__(self, missing_style_name, style_set):
        super(TempleteStyleNotFound, self).__init__(
            "The style '%s' has not been declared. Avaliable styles are: %s)"
            % (missing_style_name, style_set.names)
        )


class NoTableColumns(TableSheetException):
    def __init__(self, table_sheet):
        super(NoTableColumns, self).__init__(
            "The TableSheet '%s' has no columns. Declare atleast one."
            % table_sheet.sheetname
        )


class HeadersNotFound(TableSheetException):
    def __init__(self, table_sheet):
        super(HeadersNotFound, self).__init__(
            "Header column not found on sheet '%s' either make sure that the following headers are "
            "present '%s'." % (
                table_sheet.sheetname,
                ", ".join(table_sheet.headers)
            )
        )


class MultipleFrozenColumns(TableSheetException):
    def __init__(self, table_sheet, frozen_columns):
        super(MultipleFrozenColumns, self).__init__(
            "TableSheet '%s' has more than one frozen columns. Frozen columns: %s" % (
                type(table_sheet).__name__,
                ", ".join(column.header for column in frozen_columns)
            )
        )


class CannotHideOrGroupLastColumn(TableSheetException):
    def __init__(self):
        super(CannotHideOrGroupLastColumn, self).__init__(
            "Hiding or grouping the last column when hiding all excessive columns is rendered poorly in excel."
        )


class TableSheetExceptionPolicy(Enum):
    RaiseCellException = 1
    RaiseRowException = 2
    RaiseSheetException = 3
    IgnoreRow = 4


class TableSheet(TemplatedWorksheet):
    item_class = TableColumn

    _table_name = Typed("table_name", expected_type=str, allow_none=True)

    title_style = Typed("title_style", expected_type=str, value="Title")
    description_style = Typed("description_style", expected_type=str, value="Description")

    format_as_table = Typed("format_as_header", expected_type=bool, value=True)
    freeze_header = Typed("freeze_header", expected_type=bool, value=True)
    print_title_rows = Typed("print_title_rows", expected_types=[str, bool], value=True, allow_none=True)

    freeze_column = Typed("freeze_column", expected_types=[int, bool], value=False)
    print_title_columns = Typed("print_title_columns", expected_types=[str, int, bool], value=False, allow_none=True)
    hide_excess_columns = Typed("hide_excess_columns", expected_type=bool, value=True)
    row_styles = None

    # print_setup = Typed("print_setup", expected_types=PrintPageSetup, value=None, allow_none=True)
    # fit_to_width = Typed("fit_to_width", expected_types=)

    look_for_headers = Typed("look_for_headers", expected_type=bool, value=True)
    suffix_duplicated_headers = Typed("suffix_duplicated_headers", expected_type=bool, value=True)
    exception_policy = Typed(
        "exception_policy",
        expected_type=TableSheetExceptionPolicy,
        value=TableSheetExceptionPolicy.RaiseCellException
    )

    _first_data_cell = None
    _last_data_cell = None
    _first_header_cell = None
    _last_header_cell = None
    _row_class = None
    _column_index = 1

    def __init__(self, sheetname=None, active=None, table_name=None, title_style=None, description_style=None,
                 format_as_table=None, freeze_header=None, hide_excess_columns=None, look_for_headers=None,
                 exception_policy=None, columns=None, print_title_rows=None, print_title_columns=None,
                 suffix_duplicated_headers=None, freeze_column=None, row_styles=None):
        super(TableSheet, self).__init__(sheetname=sheetname, active=active)

        self._table_name = table_name
        self.title_style = title_style
        self.description_style = description_style
        self.format_as_table = format_as_table
        self.freeze_header = freeze_header
        self.freeze_column = freeze_column
        self.hide_excess_columns = hide_excess_columns
        self.look_for_headers = look_for_headers
        self.exception_policy = exception_policy
        self.print_title_rows = print_title_rows
        self.print_title_columns = print_title_columns
        self.suffix_duplicated_headers = suffix_duplicated_headers

        self.columns = []
        self._column_headers_counter = Counter()
        for object_attribute, column in self._items.items():
            self.add_column(column, object_attribute=object_attribute)
        self.row_styles = row_styles or self.row_styles or []

        for column in columns or []:
            self.add_column(column)
        self.add_row_style(*self.row_styles)

        self._validate()

    def _validate(self):
        self._check_atleast_one_column()
        self._check_unique_column_headers()
        self._check_max_one_frozen_column()
        self._check_last_column_not_hidden_or_grouped_if_hide_excess_columns()

    def _check_atleast_one_column(self):
        if not self.columns:
            raise NoTableColumns(self)

    def _check_unique_column_headers(self):
        if len(set(column.header for column in self.columns)) < len(self.columns):
            raise ColumnHeadersNotUnique(self.columns)

    def _check_max_one_frozen_column(self):
        frozen_columns = tuple(column for column in self.columns if column.freeze)
        if len(frozen_columns) > 1:
            raise MultipleFrozenColumns(self, frozen_columns)

    def _check_last_column_not_hidden_or_grouped_if_hide_excess_columns(self):
        if self.hide_excess_columns:
            last_column = self.columns[-1]
            if last_column.hidden or last_column.group:
                raise CannotHideOrGroupLastColumn()

    def add_column(self, column, object_attribute=None):
        column.column_index = self._column_index
        self._column_index += 1

        if object_attribute and not column._object_attribute:
            column._object_attribute = object_attribute

        self.columns.append(column)
        self._row_class = None

        if self.row_styles:
            column.add_row_style(*self.row_styles)

        # Suffix duplicated column headers
        self._column_headers_counter[column.header] += 1
        if self._column_headers_counter[column.header] > 1 and self.suffix_duplicated_headers:
            column._header = "%s %d" % (column.header, self._headers[column.header])

        return column

    def add_row_style(self, *row_styles):
        for column in self.columns:
            column.add_row_style(*row_styles)

        self.row_styles.extend(row_styles)

    def write(self, objects=None, title=None, description=None, preserve=False):
        if not self.empty:
            if preserve:
                objects = chain(list(self.read()), objects)
            self.remove()

        worksheet = self.worksheet
        self.prepare_worksheet(worksheet)
        self.write_title(worksheet, title)
        self.write_description(worksheet, description)
        self.write_headers(worksheet)
        self.write_rows(worksheet, objects)
        self.post_process_worksheet(worksheet)

    def prepare_worksheet(self, worksheet):
        for column in self.columns:
            column.prepare_worksheet(worksheet)

    def write_title(self, worksheet, title=None):
        if not title:
            return

        title = WriteOnlyCell(ws=worksheet, value=title)
        self.template_styles.style_cell(title, self.title_style)

        worksheet.append((title,))

    def write_description(self, worksheet, description=None):
        if not description:
            return

        description = WriteOnlyCell(ws=worksheet, value=description)
        self.template_styles.style_cell(description, self.description_style)

        worksheet.append((description,))

    def write_headers(self, worksheet):
        headers = tuple(
            column.create_header(worksheet, self.template_styles)
            for column in self.columns
        )

        self.worksheet.append(headers)

        self._first_header_cell = headers[0]
        self._last_header_cell = headers[-1]

    def write_rows(self, worksheet, objects=None):
        self._first_data_cell = None
        cells = None
        for index, obj in enumerate(objects):
            row_type = self.row_type(obj, index)
            cells = tuple(
                column.create_cell(
                    worksheet,
                    self.template_styles,
                    column.get_value_from_object(obj, row_type=row_type),
                    row_type=row_type
                ) for column in self.columns
            )
            worksheet.append(cells)

            if not self._first_data_cell:
                self._first_data_cell = cells[0]

            for cell, column in zip(cells, self.columns):
                column.post_process_cell(worksheet, self.template_styles, cell, row_type=row_type)

        if cells:
            self._last_data_cell = cells[-1]

    def post_process_worksheet(self, worksheet):
        first_row = (self._first_data_cell or self._first_header_cell).row
        last_row = (self._last_data_cell or self._first_header_cell).row

        for column in self.columns:
            column_letter = column.column_letter

            column.post_process_worksheet(
                worksheet,
                self.template_styles,
                first_row=first_row,
                last_row=last_row,
                data_range="%s%s:%s%s" % (column_letter, first_row, column_letter, last_row)
            )

        if self.format_as_table:
            worksheet.add_table(
                Table(
                    ref="%s:%s" % (
                        self._first_header_cell.coordinate,
                        self._last_data_cell.coordinate if self._last_data_cell
                        else "{0}{1}".format(self._last_header_cell.column, self._last_header_cell.row + 1)
                    ),
                    displayName=self.table_name,
                )
            )

        # Freeze pane
        if self.freeze_header:
            row = (self._first_data_cell or self._first_header_cell).row
        else:
            row = 1
        try:
            column = next(column.column_index for column in self.columns if column.freeze)
        except StopIteration:
            column = 0
        if row + column > 1:
            worksheet.freeze_panes = worksheet["%s%s" % (get_column_letter(column+1), row)]

        # Print titles
        if self.print_title_rows:
            if type(self.print_title_rows) == str:
                print_title_rows = self.print_title_rows
            else:
                print_title_rows = "1:%d" % self._first_header_cell.row
            worksheet.print_title_rows = print_title_rows
        if self.print_title_columns:
            if type(self.print_title_columns) == str:
                print_title_columns = self.print_title_columns
            elif type(self.print_title_columns) == int:
                # Transform from zero indexed to one indexed
                print_title_columns = "1:%d" % self.print_title_columns + 1
            else:
                print_title_columns = "1:1"
            worksheet.print_title_columns = print_title_columns

        # Grouping
        groups = groupby(self.columns, lambda col: col.group)
        for columns in (list(columns) for group, columns in groups if group):
            worksheet.column_dimensions.group(
                start=columns[0].column_letter,
                end=columns[-1].column_letter,
                outline_level=1,
                hidden=columns[0].hidden
            )

        if self.hide_excess_columns:
            worksheet.column_dimensions.group(
                start=get_column_letter(len(self.columns) + 1),
                end=get_column_letter(MAX_COLUMN_INDEX + 1),
                outline_level=0,
                hidden=True
            )

    def read(self, exception_policy=None, look_for_headers=None):
        header_found = not (look_for_headers if look_for_headers is not None else self.look_for_headers)
        _exception_policy = exception_policy if exception_policy is not None else self.exception_policy

        rows = self.worksheet.__iter__()
        row_number = 0
        try:
            while not header_found:
                row_number += 1
                header_found = self._is_row_header(next(rows))

            row_exceptions = []
            while True:
                row_number += 1
                try:
                    yield self.object_from_row(next(rows), row_number, exception_policy=_exception_policy)
                except CellExceptions as e:
                    if _exception_policy.value <= TableSheetExceptionPolicy.RaiseRowException.value:
                        raise e
                    else:
                        row_exceptions.append(e)
                except IgnoreRow:
                    continue

                if row_exceptions and _exception_policy.value <= TableSheetExceptionPolicy.RaiseSheetException.value:
                    raise RowExceptions(row_exceptions)
        except StopIteration:
            pass

        if not header_found:
            raise HeadersNotFound(self)

    def _is_row_header(self, row):
        for cell, header in zip(chain(row, repeat(None)), self.headers):
            if str(cell.value) != header:
                return False
        return True

    def object_from_row(self, row, row_number, exception_policy=TableSheetExceptionPolicy.RaiseCellException):
        data = OrderedDict()
        cell_exceptions = []
        for cell, column in zip(chain(row, repeat(None)), self.columns):
            try:
                data[column.object_attribute] = column._from_excel(cell)
            except CellException as e:
                if exception_policy.value <= TableSheetExceptionPolicy.RaiseCellException.value:
                    raise e
                else:
                    cell_exceptions.append(e)

        if cell_exceptions:
            raise CellExceptions(cell_exceptions)

        # return self.row_class(**data)
        return self.create_object(row_number, **data)

    def create_object(self, row_number, **data):
        return self.row_class(**data)

    def row_type(self, object, row_number):
        return type(object)

    @property
    def table_name(self):
        if not self._table_name:
            table_name = self.sheetname
            # Remove invalid characters
            table_name = re.sub('[^0-9a-zA-Z_]', '', table_name)
            # Remove leading characters until we find a letter or underscore
            self._table_name = re.sub('^[^a-zA-Z_]+', '', table_name)

        return self._table_name

    @property
    def headers(self):
        return (column.header for column in self.columns)

    @property
    def row_class(self):
        if not self._row_class:
            self._row_class = namedtuple(
                "%sRow" % self.__class__.__name__,
                (column.object_attribute for column in self.columns)
            )
        return self._row_class

    def __iter__(self):
        return self.read()

