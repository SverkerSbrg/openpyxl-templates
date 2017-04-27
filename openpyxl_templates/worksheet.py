from collections import deque

from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Protection
from openpyxl.styles.table import TableStyle
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.table import Table

from openpyxl_templates.exceptions import OpenpyxlTemplateCellException, CellExceptions, HeaderNotFound
from openpyxl_templates.style import SheetStyleMixin
from openpyxl_templates.utils import Typed

MAX_COLUMN_INDEX = column_index_from_string("XFD")


class ExcelRow:
    column_attrs = None
    row_number = None

    def __init__(self, column_attrs, row_number=None):
        self.column_attrs = column_attrs
        self.row_number = row_number

        for attr in self.column_attrs:
            setattr(self, attr, None)

    def __str__(self):
        return "Row %d: {%s}" % (
            self.row_number,
            ", ".join(("%s: '%s'" % (attr, getattr(self, attr)) for attr in self.column_attrs))
        )


class RowExceptionPolicy:
    IGNORE_ROW = "IGNORE_ROW"
    RAISE_EXCEPTION = "RAISE_EXCEPTION"
    RETURN_EXCEPTION = "RETURN_EXCEPTION"


class SheetTemplate(SheetStyleMixin):
    sheetname = Typed("sheetname", expected_type=str)
    title = Typed("title", expected_type=str, allow_none=True)
    description = Typed("description", expected_type=str, allow_none=True)

    columns = None
    freeze_header = Typed("freeze_header", expected_type=bool, value=True)
    hide_excess_columns = Typed("hide_excess_columns", expected_type=bool, value=True)
    format_as_table = Typed("format_as_table", expected_type=bool, value=True)
    table_style = Typed("table_style", expected_type=TableStyle, allow_none=True)

    row_exception_policy = RowExceptionPolicy.RAISE_EXCEPTION
    require_headers = Typed("require_headers", expected_type=bool, value=True)
    skip_rows = Typed("skip_rows", expected_type=int, value=0)

    sheet_protection = Typed("sheet_protection", expected_type=(bool, SheetProtection), value=False)

    def __init__(self, sheetname=None, title=None, description=None, columns=None, freeze_header=None,
                 hide_excess_columns=None, format_as_table=None, table_style=None,
                 row_exception_policy=None, require_headers=None, skip_rows=None, sheet_protection=None, **style_keys):

        super().__init__(**style_keys)

        self.sheetname = sheetname or self.sheetname
        self.title = title or self.title
        self.description = description or self.description
        self.columns = columns or self.columns or []

        self.freeze_header = freeze_header if freeze_header is not None else self.freeze_header
        self.hide_excess_columns = hide_excess_columns if hide_excess_columns is not None else self.hide_excess_columns
        self.format_as_table = format_as_table if format_as_table is not None else self.format_as_table
        self.table_style = table_style or table_style

        self.row_exception_policy = row_exception_policy or self.row_exception_policy
        self.require_headers = require_headers if require_headers is not None else self.require_headers
        self.skip_rows = skip_rows if skip_rows is not None else self.skip_rows

        self.sheet_protection = sheet_protection if sheet_protection is not None else self.sheet_protection

        self._column_attrs = list((column.object_attr for column in self.columns))

    def read_rows(self, worksheet):
        row_number = 0

        skip_rows = self.skip_rows
        header = not self.require_headers

        for raw_row in worksheet:
            row_number += 1

            if skip_rows:
                skip_rows -= 1
                continue

            if not header:
                header = self._is_row_header(raw_row)
                continue

            row = self.create_empty_row(row_number=row_number)

            que = deque(raw_row)
            errors = []
            for column in self.columns:
                cell = que.popleft() if que else None

                try:
                    value = column.from_excel_with_blank_check(cell)
                except OpenpyxlTemplateCellException as cell_exception:
                    value = cell_exception
                    errors.append(cell_exception)

                column.set_value_to_object(row, value)

            if errors:
                if self.row_exception_policy == RowExceptionPolicy.IGNORE_ROW:
                    continue

                row_exception = CellExceptions(cell_exceptions=errors)
                if self.row_exception_policy == RowExceptionPolicy.RETURN_EXCEPTION:
                    yield row_exception
                else:
                    raise row_exception

            yield row

        if not header:
            raise HeaderNotFound(self.sheetname, [column.header for column in self.columns])

    def write_headers(self, worksheet, style_set):
        headers = []
        for column in self.columns:
            header = WriteOnlyCell(ws=worksheet, value=column.header)
            style = style_set[column.header_style]
            if style:
                header.style = style
            headers.append(header)

        worksheet.append(headers)
        first_cell = headers[0]
        last_cell = headers[-1]
        return first_cell, last_cell

    def write_rows(self, worksheet, style_set, objects):
        styles = tuple(style_set[column.row_style] for column in self.columns)
        data_validations = tuple(column.data_validation for column in self.columns)

        for data_validation in data_validations:
            if data_validation:
                worksheet.add_data_validation(data_validation)

        rows = (
            (
                column.create_cell(worksheet, obj=obj)
                for column in self.columns
            ) for obj in objects
        )

        row_count = 0
        first_cell = None
        cells = tuple()
        for row in rows:
            row_count += 1

            cells = tuple(row)
            worksheet.append(cells)

            if not first_cell:
                first_cell = cells[0]
                if self.freeze_header:
                    worksheet.freeze_panes = first_cell

            """
                Styling is separated from creation since data_validation
                must be applied after appending to worksheet
            """
            for cell, style, data_validation in zip(cells, styles, data_validations):
                if style:
                    cell.style = style
                if data_validation:
                    # print("ADDING DATA VALIDATIONS")
                    data_validation.add(cell)

        last_cell = cells[-1] if cells else None

        return first_cell, last_cell

    def style_columns(self, worksheet, style_set):
        for index, column in enumerate(self.columns):
            column_dimension = worksheet.column_dimensions[get_column_letter(index + 1)]
            if column.width is not None:
                column_dimension.width = column.width

            row_style = style_set[column.row_style]
            if row_style:
                column_dimension.style = row_style

            column_dimension.hidden = column.hidden

        if self.hide_excess_columns:
            for i in range(len(self.columns) + 1, MAX_COLUMN_INDEX + 1):
                worksheet.column_dimensions[get_column_letter(i)].hidden = True

    def create_empty_row(self, row_number):
        return ExcelRow(self._column_attrs, row_number=row_number)

    def write_title(self, worksheet, styles):
        if not self.title:
            return

        title = WriteOnlyCell(worksheet, value=self.title)
        title_style = styles[self.title_style]
        if title_style:
            title.style = title_style
        worksheet.append(self._pad_with_empty_cells(worksheet, styles, (title,)))
        if not self.description:
            worksheet.append(self._pad_with_empty_cells(worksheet, styles))

    def write_description(self, worksheet, styles):
        if self.description:
            description = WriteOnlyCell(worksheet, value=self.description)
            description_style = styles[self.description_style]
            if description_style:
                description.style = description_style
            worksheet.append(self._pad_with_empty_cells(worksheet, styles, (description,)))

        if self.title or self.description:
            worksheet.append(self._pad_with_empty_cells(worksheet, styles, None))

    def write(self, worksheet, style_set, objects):
        self.write_title(worksheet, style_set)
        self.write_description(worksheet, style_set)
        first_header, last_header = self.write_headers(worksheet, style_set)
        first_row_cell, last_row_cell = self.write_rows(worksheet, style_set, objects)
        self.style_columns(worksheet, style_set)

        if self.format_as_table:
            table = Table(
                ref="%s:%s" % (first_header.coordinate, last_row_cell.coordinate if last_row_cell else last_header.coordinate),
                displayName=self.sheetname,
                tableStyleInfo=self.table_style
            )
            worksheet.add_table(table)

        # TODO: TEST, and finish, currently not usable
        if self.sheet_protection:
            worksheet.protection = self.sheet_protection \
                if type(self.sheet_protection) == SheetProtection \
                else SheetProtection(
                    sheet=True,
                    autoFilter=False,
            )

    def _is_row_header(self, row):
        row = deque(row)
        for column in self.columns:
            value = row.popleft().value if row else None
            if value != column.header:
                return False
        return True

    def _pad_with_empty_cells(self, worksheet, style_set, row=None):
        padded_row = list(row) if row else []

        for i in range(len(padded_row), len(self.columns)):
            cell = WriteOnlyCell(ws=worksheet, value=None)
            style = style_set[self.empty_style]
            if style:
                cell.style = style
            padded_row.append(cell)

        return padded_row
