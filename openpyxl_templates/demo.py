# coding=utf-8
import random
from datetime import date, datetime
import time
from enum import Enum
from os.path import dirname, join

from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle

from openpyxl_templates.styles import DefaultStyleSet, ExtendedStyle
from openpyxl_templates.table_sheet.columns import TableColumn, ChoiceColumn, DateColumn, CharColumn, TextColumn, \
    BoolColumn, IntColumn, FloatColumn, DatetimeColumn, TimeColumn, FormulaColumn
from openpyxl_templates.table_sheet.table_sheet import TableSheet
from openpyxl_templates.templated_workbook import TemplatedWorkbook
from openpyxl_templates.utils import SolidFill

DIR = dirname(__file__)


class Sexes(Enum):
    MALE = 1
    FEMALE = 2
    OTHER = 3


class Person:
    def __init__(self, first_name, last_name, date_of_birth, sex):
        self.first_name = first_name
        self.last_name = last_name
        self.date_of_birth = date_of_birth
        self.sex = sex

    @property
    def name(self):
        return "%s %s" % (self.first_name, self.last_name)


persons = (
    Person(
        "Jane",
        "Doe",
        date(year=1983, month=3, day=10),
        Sexes.FEMALE
    ),
    Person(
        "John",
        "Doe",
        date(year=1992, month=9, day=3),
        Sexes.MALE
    ),
    Person(
        "Mickey",
        "Mouse",
        date(year=1972, month=3, day=10),
        Sexes.MALE
    ),
    Person(
        "Goofy",
        "",
        date(year=1972, month=3, day=10),
        Sexes.MALE
    ),
    Person(
        "Minnie",
        "Mouse",
        date(year=1975, month=6, day=17),
        Sexes.FEMALE
    ),
    Person(
        u"女",
        u"娲",
        date(year=1975, month=6, day=17),
        Sexes.FEMALE
    )
)


class SexColumn(ChoiceColumn):
    def __init__(self, **kwargs):
        super(SexColumn, self).__init__(
            choices=(
                (Sexes.MALE, "Male"),
                (Sexes.FEMALE, "Female"),
                (Sexes.OTHER, "Other")
            ),
            **kwargs
        )


class TemplatedPersonsSheet(TableSheet):
    first_name = TableColumn(header="First name", width=15)
    last_name = TableColumn(header="Last name", width=15)
    sex = SexColumn(header="Sex")
    date_of_birth = DateColumn(header="Date of birth")

    hide_excess_columns = False


class DemoObject:
    def __init__(self, char, text, boolean, i, f, choice, time, date, datetime):
        self.char = char
        self.text = text
        self.boolean = boolean
        self.integer = i
        self.float = f
        self.choice = choice
        self.time = time
        self.date = date
        self.datetime = datetime


def to_timestamp(date_value):
    return time.mktime(date_value.timetuple())


def demo_objects(count=100):
    from_date = to_timestamp(datetime(year=1990, month=1, day=1))
    to_date = to_timestamp(datetime(year=2020, month=12, day=31))

    dates = list(datetime.fromtimestamp(random.uniform(from_date, to_date)) for i in range(0, count))
    dates.sort()

    for i in range(0, count):
        float = random.random() * 1000
        date = dates[i]

        yield DemoObject(
            "Object %d" % (i + 1),
            "This is a text which\n will wrap by default",
            random.choice((True, False)),
            float,
            float,
            random.randint(1, 3),
            date,
            date,
            date
        )


bold_true = Rule(
    type='expression',
    dxf=DifferentialStyle(
        font=Font(bold=True)
    ),
    formula=["$C3"]
)


class ColumnDemoSheet(TableSheet):
    table_name = "ColumnDemo"

    char = CharColumn(header="CharColumn")
    text = TextColumn(header="TextColumn", freeze=True)
    boolean = BoolColumn(header="BoolColumn", row_style="Row, integer", conditional_formatting=bold_true)
    integer = IntColumn(header="IntColumn", group=True)
    float = FloatColumn(header="FloatColumn", group=True)
    datetime = DatetimeColumn(header="DatetimeColumn", group=True)
    date = DateColumn(header="DateColumn")
    time = TimeColumn(header="TimeColumn", group=True)
    formula = FormulaColumn(header="FormulaColumn", formula="=SUM(ColumnDemo[IntColumn])")


class DemoWorkbook(TemplatedWorkbook):
    timestamp = True
    persons = TemplatedPersonsSheet(sheetname="Persons", active=True)
    column_demo = ColumnDemoSheet(sheetname="Column demo")


if __name__ == "__main__":
    workbook = DemoWorkbook(
        template_styles=DefaultStyleSet(ExtendedStyle(base="Default", name="Header", fill=SolidFill("FF0000"))))
    workbook.column_demo.write(objects=list(demo_objects(100)), title="Column demo")
    workbook.persons.write(objects=persons, title="Persons")

    filename = workbook.save(join(dirname(__file__), "demo.xlsx"))

    wb = DemoWorkbook(file=filename)
    print(list(wb.persons.read()))
