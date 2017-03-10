from datetime import datetime
from os.path import dirname, join

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle

from openpyxl_templates.columns import CharColumn, IntegerColumn, ChoiceColumn, DateTimeColumn, TimeColumn
from openpyxl_templates.utils import SolidFill
from openpyxl_templates.workbook import WorkbookTemplate
from openpyxl_templates.worksheet import SheetTemplate, RowExceptionPolicy

BASE_DIR = dirname(__file__)

choices = (
    ("lägg till", "ADD"),
    ("ta bort", "REMOVE"),
    ("inget", "NOOP"),
    (None, None),
)


class TestSheet(SheetTemplate):
    sheetname = "TestSheet"
    title = "Test blad2"
    # styles = [
    #     NamedStyle("__header__", font=Font(bold=True, color="FFFFFFFF"), fill=SolidFill("5d1738"))
    # ]
    row_exception_policy = RowExceptionPolicy.RETURN_EXCEPTION

    columns = [
        CharColumn(
            object_attr="one",
            header="Röd text",
            width=15,
            # style=CellStyle(
            #     font=Font(color="FFFF0000")
            # )
        ),
        IntegerColumn(
            object_attr="three",
            header="Svart siffra",
            width=15
        ),
        IntegerColumn(
            object_attr="six",
            header="Ytterligare en siffra",
            width=15
        ),
        # EmptyColumn(
        #     header="Empty",
        #     header_style=CellStyle(fill=SolidFill("AAAAAA")),
        #     hidden=False,
        #     width=10
        # ),
        ChoiceColumn(
            object_attr="action",
            header="Aktion",
            choices=choices,
            width=30,
            default_value=""
        ),
        DateTimeColumn(
            object_attr="date",
            header="Datum",
            width=15,
        ),
        TimeColumn(
            object_attr="time",
            header="Tid",
            width=10,
        )
    ]


class TestWorkbook(WorkbookTemplate):
    sheets = [
        TestSheet()
    ]
    active_sheet = "TestSheet"


test = TestWorkbook(load_workbook(join(BASE_DIR, "test1.xlsx").replace('\\', '/'), ))

for row in test.read_rows("TestSheet"):
    print(row)


class TestObject:
    def __init__(self, one, three, six, action, date):
        self.one = one
        self.three = three
        self.six = six

        self.action = action
        self.date = date

        self.time = date.time()

    def __str__(self):
        return str({
            "one": self.one,
            "three": self.three,
            "six": self.six,
            "action": self.action,
            "date": self.date,
            "time": self.time
        })


output_workbook = Workbook()
test_output = TestWorkbook(output_workbook)
test_output.write_sheet("TestSheet", (
    TestObject(1, 2, 3, "ADD", datetime.now()),
    TestObject(2, 3, 4, "NOOP", datetime.now()),
    TestObject(3, 4, 5, "REMOVE", datetime.now()),
    TestObject(4, 5, 6, "ADD", datetime.now()),
))

# alignment = Alignment(wrap_text=False)
# border = Border()
# font = Font(size=24)
#
# for x in chain(alignment, font):
#     print(x)
#
# ws = output_workbook.create_sheet("test2")
# output_workbook.active = 3
#
# ws.column_dimensions["A"].fill = SolidFill("DDDDDD")
#
#
# TEST_COUNT = 50000
#
# def unstyled():
#     ws = output_workbook.create_sheet("unstyled")
#     fill = SolidFill("DDDDDD")
#     font = Font(size=10, bold=True)
#     alignment = Alignment(horizontal="center")
#     for i in range(0, TEST_COUNT):
#         cell = WriteOnlyCell(value=i, ws=ws)
#         ws.append((cell,))
#
#
# def styled():
#     ws = output_workbook.create_sheet("styled")
#     fill = SolidFill("ff00DD")
#     font = Font(size=10, bold=True)
#     alignment = Alignment(horizontal="center")
#     for i in range(0, TEST_COUNT):
#         cell = WriteOnlyCell(value=i, ws=ws)
#         cell.fill = fill
#         cell.font = font
#         cell.alignment = alignment
#         cell.number_format = "@"
#         ws.append((cell,))
#
#
# def namedstyle():
#     ws = output_workbook.create_sheet("namedstyle")
#     namedstyle = NamedStyle(
#         name="test_named_style",
#         fill=SolidFill("0000DD"),
#         font=Font(size=10, bold=True),
#         alignment=Alignment(horizontal="center"),
#         number_format="@"
#     )
#     for i in range(0, TEST_COUNT):
#         cell = WriteOnlyCell(value=i, ws=ws)
#         cell.style = namedstyle
#         ws.append((cell,))
#
#
# # def namedstyle_change_numberformat():
# #     ws = output_workbook.create_sheet("namedstyle_change_number_format")
# #     namedstyle = NamedStyle(
# #         name="test_named_style",
# #         fill=SolidFill("0000DD"),
# #         font=Font(size=10, bold=True),
# #         alignment=Alignment(horizontal="center"),
# #         number_format="@"
# #     )
# #     for i in range(0, 100000):
# #         cell = WriteOnlyCell(value=i, ws=ws)
# #         cell.style = namedstyle
# #         cell.number_format = "0.0"
# #         ws.append((cell,))
#
#
# print("Unstyled:", timeit(unstyled, number=1))
# print("Styled:", timeit(styled, number=1))
# print("Namedstyle:", timeit(namedstyle, number=1))
# # print("Namedstyle, change:", timeit(namedstyle_change_numberformat, number=1))
#
# print(output_workbook.named_styles, type(output_workbook.named_styles))
# for style in output_workbook._named_styles:
#     print(type(style))

output_workbook.save(join(BASE_DIR, "test_output.xlsx").replace('\\', '/'))
