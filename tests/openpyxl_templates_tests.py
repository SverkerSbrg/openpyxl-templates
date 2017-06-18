from unittest import TestCase

from openpyxl import Workbook
from openpyxl_templates.worksheet import SheetTemplate, RowExceptionPolicy

from openpyxl_templates.exceptions import BlankNotAllowed, CellExceptions
from openpyxl_templates.old.columns import CharColumn
from openpyxl_templates.old.workbook import TemplatedWorkbook, TemplatedSheet
from tests.demo import DemoTemplate


class WriteTestCase(TestCase):
    def test_write_empty(self):
        workbook = Workbook()
        template = DemoTemplate(workbook)
        template.remove_all_sheets()
        template.write_sheet("Persons", [])


class TestSheet(SheetTemplate):
    row_exception_policy = RowExceptionPolicy.RETURN_EXCEPTION
    sheetname = "test"

    columns = [
        CharColumn(object_attr="key", header="Key", allow_blank=False),
        CharColumn(object_attr="name", header="Name", allow_blank=True),
    ]


class FakeCell():
    def __init__(self, value, column, row):
        self.column = column
        self.row = row
        self.value = value

    @property
    def coordinate(self):
        return "%s%s" % (self.column, self.row)


def headers(sheet_template):
    return tuple(FakeCell(col.header, None, None) for col in sheet_template.columns)


class FakeSheet:
    def __init__(self, template, data):
        self.headers = headers(template)

        self.data = []
        for row_index, row in enumerate(data):
            cells = []
            for col_index, value in enumerate(row):
                cells.append(FakeCell(value, chr(col_index + ord('A')), row_index + 2))
            self.data.append(cells)

    def __iter__(self):
        yield self.headers

        for item in self.data:
            yield item


class ReadCase(TestCase):
    def test_do_not_allow_blank(self):
        sheet_template = TestSheet(row_exception_policy=RowExceptionPolicy.RAISE_EXCEPTION)
        fake_sheet = FakeSheet(
            sheet_template,
            (
                ("key1", "Rad 1"),
                (None, "Rad 2"),
            )
        )
        try:
            list(sheet_template.read_rows(fake_sheet))
            self.fail()
        except CellExceptions as e:
            self.assertTrue(BlankNotAllowed in (type(ce) for ce in e.cell_exceptions))


class TestTemplatedWorkbook(TemplatedWorkbook):
    sheet1 = TestSheet()

class TemplatedWorkbookTestCase(TestCase):
    def setUp(self):
        self.wb = TestTemplatedWorkbook()


    def test_identify_sheets(self):
        self.assertIsInstance(self.wb.sheet1, TemplatedSheet)
        self.assertIn(self.wb.sheet1, self.wb.templated_sheets)

# class Sheet():
#     def write(self, data):
#         # if sheet exists read first, then write
#         pass
#
#     def read(self):
#         pass
#
#
# class SheetHandler():
#     def __init__(self, sheet):
#         self.sheet = sheet
#
#     def delete(self):
#         print("Delete")
#
# class WB():
#     sheet1 = Sheet()
#     sheet2 = Sheet()
#
#     def __init__(self):
#         self.sheet1 = SheetHandler(self.sheet1)
#         self.sheet2 = SheetHandler(self.sheet2)
#
# wb = WB()
#
# wb.sheet1.delete()
#
# Workbook
# Worksheet
# load_workbook


# class Test():
#     t = 2
#     instance = None
#
#     def __new__(cls, arg):
#         print("__new__", arg)
#         if not cls.instance:
#             cls.instance = super().__new__(cls)
#         return cls.instance
#
#     def __init__(self, arg):
#         print("__init__ ", arg)
#
# t = Test(2)
# t = Test(3)